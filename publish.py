#!/usr/bin/env python3
"""
publish.py -- 銘環廠商討論網站 同步腳本

每週本社／台灣分社討論完 Excel 後，在本機執行這支腳本一次：
  1. 重新讀取「20250821_購入仕様書_完了状況一覧.xlsx」
  2. 篩選 メーカー == "銘環" 且 銘環確認 非空白、不等於 "-" 的品項
  3. 用管理番号（PS-xxxxx）在「購入仕様書」資料夾樹裡找出對應子資料夾
  4. 抓出資料夾內檔名含「購入仕様書／購入仕様図／製品外観目視検査基準／検査基準書／検査成績書」的檔案，
     同一關鍵字若同時有 PDF 跟其他格式的重複檔案，只保留 PDF
  5. 把品項資料 upsert 進 Supabase，檔案上傳到 Supabase Storage

可重複執行（idempotent）：沒有變化的品項/檔案不會重複寫入或重傳。

使用方式：
    python publish.py            # 正式執行
    python publish.py --dry-run  # 只顯示會做什麼，不連線 Supabase

環境變數（可寫在同目錄的 .env，或用系統環境變數）：
    SUPABASE_URL              例如 https://xxxxx.supabase.co
    SUPABASE_SERVICE_ROLE_KEY Supabase 專案的 service_role key（切勿外流／勿提交進版本控制）
    EXCEL_PATH                預設見下方 DEFAULT_EXCEL_PATH
    BASE_FOLDER               預設見下方 DEFAULT_BASE_FOLDER
"""
import argparse
import glob
import hashlib
import mimetypes
import os
import sys
from pathlib import Path

import openpyxl
import requests

# ---------------------------------------------------------------------------
# 預設路徑（依實際環境調整，或用環境變數 / .env 覆蓋）
# ---------------------------------------------------------------------------
DEFAULT_EXCEL_PATH = (
    r"C:\Users\q5695\株式会社ハヤカワカンパニー\HKC→HTW_小牧検査移管 - 文件"
    r"\購入仕様書\検査移管資料\20250821_購入仕様書_完了状況一覧.xlsx"
)
DEFAULT_BASE_FOLDER = (
    r"C:\Users\q5695\株式会社ハヤカワカンパニー\HKC→HTW_小牧検査移管 - 文件"
    r"\購入仕様書\購入仕様書"
)
SHEET_NAME = "Sheet1"
HEADER_ROW = 5
DATA_START_ROW = 6

COL_KANRI = 5   # E: 管理番号
COL_KOKYAKU = 6  # F: 顧客
COL_MAKER = 7    # G: メーカー
COL_HINBAN = 8   # H: 品番
COL_HINMEI = 9   # I: 品名
COL_MEIKAN = 17  # Q: 銘環確認

TARGET_MAKER = "銘環"
FILE_KEYWORDS = ["購入仕様書", "購入仕様図", "製品外観目視検査基準", "検査基準書", "検査成績書"]

BUCKET = "item-files"


def safe_storage_path(item_id: str, filename: str) -> str:
    """Supabase Storage の object key は日本語などの非 ASCII 文字を受け付けないため、
    ファイル名のハッシュ値で ASCII セーフなキーを作る（元のファイル名は item_files.filename に保存し、
    ダウンロード時は download 属性で元のファイル名を復元する）。"""
    ext = Path(filename).suffix
    digest = hashlib.sha1(filename.encode("utf-8")).hexdigest()[:20]
    return f"{item_id}/{digest}{ext}"


# ---------------------------------------------------------------------------
# 小工具：不依賴 python-dotenv，自己讀 .env
# ---------------------------------------------------------------------------
def load_dotenv(path: Path):
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip().strip('"').strip("'")
        os.environ.setdefault(key, value)


def select_matched_files(files):
    """依 FILE_KEYWORDS 逐一分組（每個檔名只算進第一個符合的關鍵字），
    同一個關鍵字底下如果同時有 PDF 跟其他格式（Word／Excel）的重複檔案，
    只保留 PDF；沒有 PDF 的話才保留原本的檔案（例如分頁的多個 PDF 會全部保留）。"""
    groups = {k: [] for k in FILE_KEYWORDS}
    for f in files:
        for k in FILE_KEYWORDS:
            if k in f:
                groups[k].append(f)
                break

    selected = []
    for k in FILE_KEYWORDS:
        group = groups[k]
        if not group:
            continue
        pdfs = [f for f in group if f.lower().endswith(".pdf")]
        selected.extend(pdfs if pdfs else group)
    return selected


# ---------------------------------------------------------------------------
# 資料擷取：讀 Excel + 比對資料夾/檔案（邏輯已於分析階段驗證過）
# ---------------------------------------------------------------------------
def extract_items(excel_path: Path, base_folder: Path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[SHEET_NAME]

    items = []
    skipped_no_folder = []

    for r in range(DATA_START_ROW, ws.max_row + 1):
        maker = ws.cell(row=r, column=COL_MAKER).value
        if maker != TARGET_MAKER:
            continue

        meikan = ws.cell(row=r, column=COL_MEIKAN).value
        meikan_str = str(meikan).strip() if meikan is not None else ""
        if meikan_str == "" or meikan_str == "-":
            continue

        kanri = ws.cell(row=r, column=COL_KANRI).value
        if not kanri or not str(kanri).startswith("PS-"):
            skipped_no_folder.append((r, kanri, "沒有有效的管理番号"))
            continue
        kanri = str(kanri).strip()

        matches = glob.glob(str(base_folder / "*" / f"{kanri}_*"))
        matches = [m for m in matches if os.path.isdir(m)]
        if not matches:
            skipped_no_folder.append((r, kanri, "找不到對應資料夾"))
            continue
        folder = Path(matches[0])

        candidate_files = [
            f for f in os.listdir(folder)
            if os.path.isfile(folder / f) and any(k in f for k in FILE_KEYWORDS)
        ]
        matched_files = select_matched_files(candidate_files)

        items.append({
            "id": kanri,
            "customer": ws.cell(row=r, column=COL_KOKYAKU).value,
            "maker": maker,
            "part_no": ws.cell(row=r, column=COL_HINBAN).value,
            "part_name": ws.cell(row=r, column=COL_HINMEI).value,
            "meikan_note": meikan_str,
            "folder_path": str(folder.relative_to(base_folder)),
            "folder_abs": folder,
            "files": matched_files,
        })

    return items, skipped_no_folder


# ---------------------------------------------------------------------------
# Supabase REST / Storage 呼叫（用 requests 直打 API，不需額外套件）
# ---------------------------------------------------------------------------
class SupabaseClient:
    def __init__(self, url: str, service_role_key: str):
        self.url = url.rstrip("/")
        self.key = service_role_key
        self.headers = {
            "apikey": service_role_key,
            "Authorization": f"Bearer {service_role_key}",
        }

    def upsert_item(self, item: dict):
        payload = {
            "id": item["id"],
            "customer": item["customer"],
            "maker": item["maker"],
            "part_no": item["part_no"],
            "part_name": item["part_name"],
            "meikan_note": item["meikan_note"],
            "folder_path": item["folder_path"],
            "updated_at": "now()",
        }
        resp = requests.post(
            f"{self.url}/rest/v1/items",
            headers={
                **self.headers,
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
            params={"on_conflict": "id"},
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()

    def list_item_files(self, item_id: str):
        resp = requests.get(
            f"{self.url}/rest/v1/item_files",
            headers=self.headers,
            params={"item_id": f"eq.{item_id}", "select": "id,filename,storage_path,size_bytes"},
            timeout=30,
        )
        resp.raise_for_status()
        return {row["filename"]: row for row in resp.json()}

    def upsert_item_file(self, item_id: str, filename: str, storage_path: str, size_bytes: int):
        payload = {
            "item_id": item_id,
            "filename": filename,
            "storage_path": storage_path,
            "size_bytes": size_bytes,
            "updated_at": "now()",
        }
        resp = requests.post(
            f"{self.url}/rest/v1/item_files",
            headers={
                **self.headers,
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
            params={"on_conflict": "item_id,filename"},
            json=payload,
            timeout=30,
        )
        resp.raise_for_status()

    def delete_item_file_row(self, row_id: int):
        resp = requests.delete(
            f"{self.url}/rest/v1/item_files",
            headers=self.headers,
            params={"id": f"eq.{row_id}"},
            timeout=30,
        )
        resp.raise_for_status()

    def upload_file(self, storage_path: str, local_path: Path):
        content_type, _ = mimetypes.guess_type(str(local_path))
        with open(local_path, "rb") as f:
            data = f.read()
        resp = requests.post(
            f"{self.url}/storage/v1/object/{BUCKET}/{storage_path}",
            headers={
                **self.headers,
                "Content-Type": content_type or "application/octet-stream",
                "x-upsert": "true",
            },
            data=data,
            timeout=120,
        )
        resp.raise_for_status()

    def delete_storage_object(self, storage_path: str):
        resp = requests.delete(
            f"{self.url}/storage/v1/object/{BUCKET}/{storage_path}",
            headers=self.headers,
            timeout=30,
        )
        # 404 表示本來就不存在，忽略即可
        if resp.status_code not in (200, 404):
            resp.raise_for_status()


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    # Windows のコンソールが cp932 などの場合に日本語/中国語の print で落ちないようにする
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(description="同步銘環品項與檔案到 Supabase")
    parser.add_argument("--excel", default=None, help="Excel 檔案路徑（覆蓋預設值）")
    parser.add_argument("--base-dir", default=None, help="購入仕様書 資料夾路徑（覆蓋預設值）")
    parser.add_argument("--dry-run", action="store_true", help="只顯示會做什麼，不連線 Supabase")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    load_dotenv(script_dir / ".env")

    excel_path = Path(args.excel or os.environ.get("EXCEL_PATH", DEFAULT_EXCEL_PATH))
    base_folder = Path(args.base_dir or os.environ.get("BASE_FOLDER", DEFAULT_BASE_FOLDER))

    if not excel_path.exists():
        sys.exit(f"[錯誤] 找不到 Excel 檔案：{excel_path}")
    if not base_folder.exists():
        sys.exit(f"[錯誤] 找不到資料夾：{base_folder}")

    print(f"讀取 Excel：{excel_path}")
    print(f"掃描資料夾：{base_folder}")
    items, skipped = extract_items(excel_path, base_folder)
    print(f"符合條件品項：{len(items)} 筆")
    if skipped:
        print(f"略過（無法對應資料夾/管理番号）：{len(skipped)} 筆")
        for r, kanri, reason in skipped:
            print(f"  - Excel row {r}（管理番号={kanri}）：{reason}")

    if args.dry_run:
        print("\n--dry-run 模式，以下為預覽，不會寫入 Supabase：\n")
        for item in items:
            print(f"[{item['id']}] {item['customer']} / {item['part_no']} / {item['part_name']}")
            print(f"    銘環確認: {item['meikan_note']}")
            print(f"    資料夾: {item['folder_path']}")
            for fn in item["files"]:
                print(f"    - {fn}")
        return

    supabase_url = os.environ.get("SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not supabase_url or not service_key:
        sys.exit(
            "[錯誤] 缺少 SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY。\n"
            "請在 publish.py 同目錄建立 .env（可參考 .env.example）或設定環境變數。"
        )

    client = SupabaseClient(supabase_url, service_key)

    new_count = updated_count = uploaded_count = skipped_upload_count = removed_count = 0

    for item in items:
        client.upsert_item(item)
        updated_count += 1

        existing_files = client.list_item_files(item["id"])
        current_filenames = set(item["files"])

        for filename in item["files"]:
            local_path = item["folder_abs"] / filename
            size_bytes = local_path.stat().st_size
            existing = existing_files.get(filename)
            if existing and existing.get("size_bytes") == size_bytes:
                skipped_upload_count += 1
                continue

            storage_path = safe_storage_path(item["id"], filename)
            client.upload_file(storage_path, local_path)
            client.upsert_item_file(item["id"], filename, storage_path, size_bytes)
            uploaded_count += 1
            print(f"  上傳：[{item['id']}] {filename}")

        # 移除已經不在資料夾內（例如改版後檔名變更）的舊檔案紀錄
        for filename, row in existing_files.items():
            if filename not in current_filenames:
                client.delete_storage_object(row["storage_path"])
                client.delete_item_file_row(row["id"])
                removed_count += 1
                print(f"  移除舊檔：[{item['id']}] {filename}")

    print("\n完成：")
    print(f"  品項 upsert：{updated_count} 筆")
    print(f"  檔案上傳：{uploaded_count} 個（略過未變更 {skipped_upload_count} 個）")
    print(f"  移除舊檔：{removed_count} 個")


if __name__ == "__main__":
    main()
